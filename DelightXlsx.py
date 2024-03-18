# from openpyxl import load_workbook, Workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell, xl_col_to_name

global countTestItems
global stressTestTimes


class XlsxManager(object):

    def __init__(self):
        pass

    def clone_by_weigh(self, writer, shape, sheet_namaiwa, new_excel_filename, weigh):
        try:
            # Get the xlsxwriter objects from the dataframe writer object.
            workbook = writer.book
            worksheet = sheet_namaiwa
            # workbook_clone = xlsxwriter.Workbook(f'{new_excel_filename}')
            workbook_clone = workbook.add_worksheet(new_excel_filename)
            # Set a formula for each cell in the row
            df_row_num = shape[0]  # Row number
            df_col_num = shape[1]  # Col number
            start_col = 7  # Start column number (after the inserted column)
            end_col = df_col_num  # End column number (for example)
            # format1 = workbook.add_format({"bg_color": "#FFC7CE", "font_color": "#9C0006"})
            # --------- complement LSL, USL, AVG, STD, CPK, cols ---------
            for r, row in enumerate(writer.values):
                for c, values in enumerate(row[:8]):  # until GGR columns
                    workbook_clone.write(r, c, values)

            # --------- clone weigh data to other sheets ---------
            for row in range(2, df_row_num + 2):
                for col_num in range(start_col, end_col):
                    col_alphabet_head = xl_col_to_name(col_num)  # start from H2
                    col_alphabet_tail = xl_col_to_name(col_num + weigh - 1)  # start from H2
                    formula = f'=IFERROR(AVERAGE({sheet_namaiwa}!{col_alphabet_head}{row}:{col_alphabet_tail}{row}), "N/A")'
                    workbook_clone.write_formula(f'{col_alphabet_head}{row}', formula)
            # --------- overlap the calculation misplace part ---------
            for row in range(2, df_row_num + 2):
                for col_num in range(0, weigh - 1):
                    col_alphabet_head = xl_col_to_name(col_num)  # start from H2
                    col_alphabet_tail0 = xl_col_to_name(end_col - weigh + 2 + col_num)  # start from H2
                    col_alphabet_tail_end = xl_col_to_name(end_col)  # start from H2
                    formula = f'=AVERAGE({sheet_namaiwa}!{col_alphabet_head}{row}, {sheet_namaiwa}!{col_alphabet_tail0}{row}:{col_alphabet_tail_end}{row})'
                    # workbook_clone.write_formula(f'{col_alphabet_tail0}{row}', formula)
                    if weigh == 2:
                        workbook_clone.write_formula(f'{col_alphabet_tail0}{row}', formula)
                        # workbook_clone.write_formula(f'{col_alphabet_tail_end}{end_col}', formula)
                    elif weigh == 3:
                        # workbook_clone.write_formula(f'{col_alphabet_tail_end}{df_col_num - weigh}', formula)
                        workbook_clone.write_formula(f'{col_alphabet_tail0}{row}', formula)
                    elif weigh == 4:
                        workbook_clone.write_formula(f'{col_alphabet_tail0}{row}', formula)
                    elif weigh == 5:
                        workbook_clone.write_formula(f'{col_alphabet_tail0}{row}', formula)


            # Close the workbook at end of weigh loop
            # worksheet.conditional_format(f'F1:F{df_row_num}', {"type": "cell", "criteria": "<", "value": 1.33, "format": format1})
            # workbook.close()
        except Exception as e:
            raise "cooking_CPK$NG >>> " + str(e.args)

    def cooking_xCPK(self, writer, shape):
        try:
            # Get the xlsxwriter objects from the dataframe writer object.
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            # Set a formula for each cell in the row
            df_row_num = shape[0]  # Row number
            df_col_num = shape[1]  # Col number
            start_col = 4  # Start column number (after the inserted column)
            end_col = 6  # End column number (for example)
            corresponding_column = xl_col_to_name(df_col_num - 5)  # col_alphabet = 'AO'  # annotate upon column
            format1 = workbook.add_format({"bg_color": "#FFC7CE", "font_color": "#9C0006"})
            # get rid of extra added columns: AVG, STD, CPK, LSL, USL
            for row in range(1, df_row_num + 1):
                for col_num in range(start_col, end_col):
                    formula = f'=IFERROR(AVERAGE(G{row + 1}:{corresponding_column}{row + 1}), "N/A")'
                    worksheet.write_formula(f'D{row + 1}', formula)
            # ---------------------------------------------------------------------------
            for row in range(1, df_row_num + 1):
                for col_num in range(start_col, end_col):
                    formula = f'=IFERROR(STDEV(G{row + 1}:{corresponding_column}{row + 1}), "N/A")'
                    worksheet.write_formula(f'E{row + 1}', formula)
            # ---------------------------------------------------------------------------
            for row in range(1, df_row_num + 1):
                for col_num in range(start_col, end_col):
                    formula = f'=IFERROR(MIN((C{row + 1} - D{row + 1}) / (3 * E{row + 1}),(D{row + 1} - B{row + 1})/ (3 * E{row + 1})),"N/A")'
                    worksheet.write_formula(f'F{row + 1}', formula)
                    # worksheet.write_formula(cell_ref, formula)
            # Close the workbook
            # Saving the modified Excel file in default (that is Excel 2003) format
            # Write a conditional format over a range.
            worksheet.conditional_format(
                f'F1:F{df_row_num}', {"type": "cell", "criteria": "<", "value": 1.33, "format": format1}
            )
            workbook.close()
        except Exception as e:
            raise "cooking_CPK$NG >>> " + str(e.args)

    def df_constructor(self, df, df_weigh):
        try:
            pass
        except Exception as e:
            raise ""

    # def grr_packingXlsx(self, *args):
    #     data_path, grr_lst, avg_weigh = args
    #     current_time = datetime.datetime.now()
    #     formatted_time = current_time.strftime(f'%H%M%S')
    #
    #     try:
    #         summary_path = os.path.join(self.pwd, "Summary")
    #         wb = openpyxl.load_workbook(data_path, data_only=False)
    #         # Get the active sheet
    #         sheet = wb.active
    #         grr_lst = [np.nan if val == 'NAN' else str(val) for val in grr_lst]
    #         # num_lst = [(str(i + 1)) for i in range(sheet.max_row)]
    #         sheet.insert_cols(7)
    #         sheet.cell(row=1, column=7, value="GRR")
    #         for row_num, value in enumerate(grr_lst, start=2):
    #             sheet.cell(row=row_num, column=7, value=value)
    #         red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    #         for cell in sheet['G']:
    #             cell_letter = cell.coordinate
    #             if cell_letter != 'G1' and float(cell.value) > 30:
    #                 sheet.conditional_formatting.add(cell_letter, CellIsRule(operator='greaterThan', formula=['30'], fill=red_fill))
    #
    #         summary_file_path = os.path.join(summary_path, f'GRR_avg{avg_weigh}_{formatted_time}.xlsx')
    #         wb.save(summary_file_path)
    #         # df_pack.to_excel(summary_file_path, engine='xlsxwriter')
    #     except Exception as e:
    #         raise "grr_packing NG >>>" + str(e.args)

    # def Openpyxl_cooking_CPK(self, df):
    #
    #     try:
    # Load the workbook
    # wb = Workbook()
    # sheet = wb.active
    # Write the DataFrame to the Excel sheet
    # for r in dataframe_to_rows(df, index=True, header=True):
    #     sheet.append(r)
    # Set values for cells in the row
    # for col_num in range(5810):  # Assuming 10 columns
    #     sheet.cell(row=1, column=col_num, value=col_num)

    # Set a formula for each cell in the row
    #     sheet.insert_cols(2)
    #     sheet.insert_cols(2)
    #     sheet.insert_cols(2)
    #     for row in range(5810):
    #         cell = sheet.cell(row=row + 1, column=4)
    #         cell.formula = f'=IFERROR(AVERAGE(G{row + 2}:AP{row + 2}), "N/A")'
    #         cell = sheet.cell(row=row + 1, column=5)
    #         cell.formula = f'=IFERROR(STDEV(G{row + 2}:AP{row + 2}), "N/A")'
    #         cell = sheet.cell(row=row + 1, column=6)
    #         cell.formula = f'=IFERROR(MIN((C{row + 2} - D{row + 2}) / (3 * E{row + 2}),(D{row + 2} - B{row + 2})/ (3 * E{row + 2})),"N/A")'
    #
    #     # Save the workbook
    #     wb.save('example.xlsx')
    # except Exception as e:
    #     raise "cooking_CPK$NG >>> " + str(e.args)

    # @staticmethod
    # def highlight_NG(df):
    #     try:
    #         # Create a new Excel workbook and select the active sheet
    #         wb = Workbook()
    #         ws = wb.active
    #
    #         # Write the DataFrame to the Excel sheet
    #         for r in dataframe_to_rows(df, index=True, header=True):
    #             ws.append(r)
    #
    #         # Apply conditional formatting to highlight values greater than 30
    #         # for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    #         #     for cell in row:
    #         for row in range(2, ws.max_row + 1):
    #             for col in range(2, ws.max_column - 6):  # highlight without AVG cols
    #                 cell = ws.cell(row=row, column=col)
    #                 if cell.value is not None and cell.value > 30:
    #                     cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    #
    #         # Save the workbook
    #         wb.save('highlighted_values.xlsx')
    #     except Exception as e:
    #         raise "grr_summary NG >>> " + str(e.args)

    # @staticmethod
    # def digest_by_openpyxl(path):
    #     try:
    #         wb = load_workbook(path)
    #
    #         # Iterate over each sheet in the workbook
    #         for sheet in wb.sheetnames:
    #             # Create a new CSV file for each sheet
    #             with open(f'{sheet}.csv', 'w', newline='') as csvfile:
    #                 csvwriter = csv.writer(csvfile)
    #
    #                 # Copy the data from the original sheet to the CSV file
    #                 for row in wb[sheet].iter_rows(values_only=True):
    #                     csvwriter.writerow(row)
    #
    #                 # Insert a row with only one value in the first cell
    #                 csvwriter.writerow(['Your Value'])  # Insert your desired value here
    #     except Exception as e:
    #         raise "digest_by_openpyxl NG >>>" + str(e.args)
