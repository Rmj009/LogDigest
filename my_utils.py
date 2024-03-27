import io
import os
import re
import datetime
import numpy as np
import pandas as pd
import openpyxl

import xlsxwriter
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from DelightXlsx import XlsxManager
from Gage import Gage
from mainViz import RfVisualize


class Digest_utils:
    def __init__(self, pwd):
        self.pwd = pwd
        self.output_path = os.path.join(pwd, "Result")
        self.xlsxInstance = XlsxManager()
        # self.proj_alias = project_name

    def _plt_chart(self, *args):
        try:
            data_path, avg_weigh = args
            chart_instance = RfVisualize(self.pwd)
            df = pd.read_excel(data_path, index_col=0, header=0, engine='openpyxl')  # sheet_name=f'GRR{avg_weigh}',
            df = df[(df['USL']) != 0 | (df['LSL'] != 0)]  # drop both USL LSL => zero imply GRR => NAN
            lsl_lst = df['LSL']
            usl_lst = df['USL']
            df = df.iloc[:, 6:]  # acquire data except avg, usl, lsl..,
            for i, row in enumerate(df.iterrows(), start=2):
                # rowData = row[1][6:]
                lsl = lsl_lst.iloc[i - 2]
                usl = usl_lst.iloc[i - 2]
                # lsl = df.iloc[i - 2]['LSL']
                # usl = df.iloc[i - 2]['USL']
                cpk, cpl, cpu = self._cpk_cooking(row, lsl, usl)
                if not np.isnan(cpk) or not np.isinf(cpk):
                    # chart_instance.shewhart(row, avg_weigh, cpk)
                    chart_instance.CPK_Chart(row, lsl, usl, cpk, cpl, cpu)
        except Exception as e:
            print("_plt_chart$exception >>> " + str(e.args))
            raise str(__name__) + f'$NG{e.args}'

    def _cpk_cooking(self, *args):
        try:
            data, lsl, usl = args
            testdata = data[1]
            mean = np.mean(testdata)
            stdev = np.std(testdata)
            cpl = (usl - mean) / (3 * stdev)
            cpu = (mean - lsl) / (3 * stdev)
            cpk = min(cpl, cpu)
        except Exception as e:
            print("_cpk_cooking$exception >>> " + str(e.args))
            return np.nan
        return cpk, cpl, cpu

    def grr_calculation(self, data: np.array, USL, LSL) -> str:
        """
        loop columns to calc
        """
        try:
            USL = float(USL)
            LSL = float(LSL)
            # df_testItem = pd.DataFrame(np.array(df.iloc[:, col_nth]).reshape(10, 9))
            df_testItem = pd.DataFrame(data)
            df_testItem.columns = ['A', 'A', 'A', 'B', 'B', 'B', 'C', 'C', 'C']
            df_testItem.index = [i for i in range(10)]
            A_op = df_testItem.iloc[:, :3].values
            B_op = df_testItem.iloc[:, 3:6].values
            C_op = df_testItem.iloc[:, 6:].values
            # a = np.all(grr_data == np.swapaxes(df_testItem.values, 1, 0))
            grr_data = [A_op, B_op, C_op]
            grr_instance = Gage(grr_data, LSL, USL)
            grr_value = grr_instance.cooking_grr()
        except Exception as e:
            print("grr_calculation exception >>> " + str(e.args))
            return np.nan
        return grr_value

    def grr_cooking(self, df, csv_path, avg_weigh):  # vertical GRR calc
        """
        split dataframe into A,B,C blocks
        :return:
        """
        grr_lst = [f'GRR{avg_weigh}']
        avg_lst = [f'AVG{avg_weigh}']
        LSL = df['LSL']
        USL = df['USL']
        try:
            df = pd.read_csv(f'{csv_path}', header=0, index_col=0)
            spec_array = df.iloc[[0, 1]].values[:, 1:]
            df = df.iloc[2:, :-1]  # ignore last column
            for i in range(df.shape[1] - 1):
                select_arr_ith = np.array(df.iloc[:, i + 1]).reshape(10, 9)
                avg_lst.append(np.mean(select_arr_ith))
                grr_lst.append(self.grr_calculation(select_arr_ith, USL[i], LSL[i]))
                # grr_lst.append(self.grr_calculation(select_arr_ith, spec_array[:, i]))
            # df_result = pd.concat([df.iloc[:3], result, df.iloc[3:]]).reset_index(drop=True)

        except Exception as e:
            raise "grr_cooking NG >>> " + str(e.args)
        return self.grr_packingCsv(self.pwd, csv_path, avg_lst, grr_lst, avg_weigh)

    def grr_roasting(self, df):  # vertical GRR calc
        try:
            grr_lst = []
            # df = pd.read_excel(data_path, index_col=0, header=0, engine='openpyxl')
            # gg = df.eq(np.nan)
            # columns_to_drop = df.dtypes[df.dtypes == np.dtype(None)].index
            # df.mask(df.astype(None).eq(None)).dropna()
            # df.dropna(axis=0, how='all')
            LSL = df['LSL']
            USL = df['USL']
            for i, row in df.iterrows():
                select_arr_ith = np.array(row[5:95]).reshape(10, 9)
                grr_lst.append(self.grr_calculation(select_arr_ith, USL[i], LSL[i]))
        except IOError as ioe:
            print("Permission denied, because xlsx is opened! >>" + str(ioe.args))
        except Exception as e:
            print("grr_roasting$NG >>> " + str(e.args))
            raise "grr_roasting$NG >>> " + str(e.args)
        return grr_lst

    def grr_packingXlsx(self, *args):
        grr_lst, avg_weigh = args
        current_time = datetime.datetime.now()
        formatted_time = current_time.strftime(f'%H%M%S')
        ft = Font(color="FF0000", bold=True)
        files = os.listdir(self.output_path)
        # Sort the files based on modification time (newest first)
        files.sort(key=lambda x: os.path.getmtime(os.path.join(self.output_path, x)), reverse=True)
        cpk_xlsx_path = os.path.join(self.output_path, files[0])
        try:
            wb = openpyxl.load_workbook(filename=cpk_xlsx_path, data_only=False)
            # Get the active sheet
            sheet = wb.active
            grr_lst = [np.nan if val == np.nan else str(val) for val in grr_lst]
            sheet.insert_cols(7)
            GRR_cell = sheet.cell(row=1, column=7, value="GRR")
            GRR_cell.font = ft
            for row_num, value in enumerate(grr_lst, start=2):
                sheet.cell(row=row_num, column=7, value=value)
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            for cell in sheet['G']:
                cell_letter = cell.coordinate
                if cell_letter != 'G1' and float(cell.value) > 30:
                    sheet.conditional_formatting.add(cell_letter, CellIsRule(operator='greaterThan', formula=['30'], fill=red_fill))

            result_file_path = os.path.join(os.path.join(self.pwd, "Result"), f'Result_avg_{formatted_time}.xlsx')
            wb.save(result_file_path)
            # df_pack.to_excel(summary_file_path, engine='xlsxwriter')
        except Exception as e:
            raise "grr_packing NG >>>" + str(e.args)

    def grr_packingCsv(self, *args):
        pwd, csv_path, avg_lst, grr_lst, avg_weigh = args
        current_time = datetime.datetime.now()
        formatted_time = current_time.strftime(f'%H%M%S')
        try:
            summary_path = os.path.join(pwd, "Summary")
            df_pack = pd.read_csv(f'{csv_path}', header=0)
            df_pack = df_pack.iloc[:, :-1]  # remove last col fixture
            # grr_lst = [np.nan if val == 'NAN' else float(val) for val in grr_lst]
            grr_lst.insert(0, f'GRR{avg_weigh}')  # GRR as row name
            avg_lst.insert(0, f'AVG{avg_weigh}')  # AVG as row name
            #  ------------- add AVG attribute -------------
            df_pack.index = df_pack.index[:2].tolist() + (df_pack.index[2:] + 2).tolist()
            df_pack.loc[3] = np.array(avg_lst)
            df_pack.loc[2] = np.array(grr_lst)
            #  ------------- add GRR attribute -------------
            # df_pack.index = df_pack.index[:2].tolist() + (df_pack.index[2:] + 1).tolist()
            # df_pack.loc[2] = np.array(grr_lst)
            df_pack = df_pack.sort_index()
            summary_file_path = os.path.join(summary_path, f'GRR_avg{avg_weigh}_{formatted_time}.csv')
            df_pack.to_csv(summary_file_path, sep=',')
        except Exception as e:
            raise "grr_packing NG >>>" + str(e.args)

    def grr_data_digest(self, *args, **kwargs) -> np.array:
        filepath = args

        try:
            csv_data_path = os.path.join(self.pwd, "DataLog")
            if not os.path.exists(csv_data_path):
                os.mkdir(csv_data_path)
                print("Directory '% s' created" % csv_data_path)
            else:
                all_csv = os.listdir(csv_data_path)
                for file in all_csv:
                    if file.endswith(".csv"):
                        file_path = os.path.join(csv_data_path, file)
                        os.remove(file_path)
                        print(f'remove csv >>>{file_path}')
            IsLastCol_Literal = True
            if IsLastCol_Literal:
                self.digest_csv(filepath, csv_data_path, 3)
            # df_result.to_csv(f'weight{window}.csv', sep=',')
        except Exception as e:
            raise "grr_data_digest NG >>>" + str(e.args)
        # return df_spec_array  # str(df.shape)
        # return count_total_csv

    def digest_csv(self, *args):
        csv_path, csv_data_path, weighted = args
        """
        weighted average by rolling function
        """
        mockup_col_name = []
        try:
            df = pd.read_csv(f'{csv_path}', header=0, index_col=0)
            mockup_col_name = df.columns[1:-1]
            df_values = df.iloc[2: df.shape[0], 1: df.shape[1] - 1]  # all values
            for w in range(weighted):
                nest_lst = []
                # for row in range(loop_row):  # dut_test_times per dut = 9     loop_row = df.shape[0] - 3
                for col in range(df_values.shape[1]):
                    arr_weight = df_values.iloc[:, col].rolling(w + 2, min_periods=w + 1, center=True).mean().values
                    # nest_lst.append(arr_weight)
                    nest_lst.append(arr_weight)
                csv_file_name = f'GRR{w + 1}.csv'
                csv_path_ = os.path.join(csv_data_path, csv_file_name)
                # flat_lst = [item for sublist in arr_lst for item in sublist]
                flat_lst = [arr for arr in nest_lst]
                df_result = pd.concat([pd.DataFrame(arr) for arr in flat_lst], axis=1)
                # df_result = pd.merge([pd.DataFrame(arr) for arr in flat_lst], on='left')
                df_result.columns = mockup_col_name
                df_result.to_csv(csv_path_, index=False)
            # print(f'Sheet "{sheet_name}" saved as "{csv_file_name}"')
            # for i in range(df.shape[1] - 1):
            #     select_arr_ith = np.array(df.iloc[:, i + 1]).reshape(10, 9)
            #     df_testItem = pd.DataFrame(select_arr_ith)
            #     df_testItem = df_testItem.T
            #     df_testItem.index = ['A', 'A', 'A', 'B', 'B', 'B', 'C', 'C', 'C']
            #     df_testItem.columns = [i for i in range(10)]
            #     grr_data = pd.DataFrame(df_testItem.values.reshape(3, 10, 3))
            #     grr_data.loc['A'].apply(lambda x: x.rolling(window=2).mean())

        except Exception as e:
            raise "xlsx file NG >>>" + str(e.args)

    def digest_xlsx(self, *args):
        result_xlsx_filename = args[0]
        current_time = datetime.datetime.now()
        formatted_time = current_time.strftime(f'%H%M%S')
        try:
            df = pd.read_excel(result_xlsx_filename, index_col=0, header=0, engine='openpyxl')
            # df = df[(df['USL']) != 0 | (df['LSL'] != 0)]  # drop both USL LSL => zero imply GRR => NAN
            # xls = pd.ExcelFile(result_xlsx_filename, engine='openpyxl')  # "Weighed_result.xlsx"
            # sheet_names_inXls = [ n.parse(sheet_name=name) for n in xls.sheet_names]
            # xls = pd.ExcelWriter("tryMakingGRR.xlsx", engine='xlsxwriter')  # "Weighed_result.xlsx"
            # sheet_name = xls.sheet_names  # barely one sheet literally
            # df = pd.read_excel(xls, sheet_name=xls.sheet_names[0])
            # workbook = xlsxwriter.Workbook(result_xlsx_filename)
            # data_only=True will remove the formula
            wb = openpyxl.load_workbook(result_xlsx_filename, data_only=False)
            for weigh in range(2, 6):
                ws = wb.create_sheet(f'GRR{weigh}', weigh - 1)
                df_GRR = self.grr_weigh_pack(df, weigh)
                for r in dataframe_to_rows(df_GRR, index=True, header=True):
                    ws.append(r)
                ws.delete_rows(2)
                # self.xlsxInstance.openpyxl_cooking_CPK(sheet=ws, max_cols=df_GRR.shape[1] + 1)
                last_col_literal = get_column_letter(df_GRR.shape[1] + 1)
                for i, cell in enumerate(ws['D'], start=1):  # start from 1 bcoz the skip the first cell => AVG
                    formula = f'=IFERROR(AVERAGE(H{i}:{last_col_literal}{i}), "N/A")'
                    cell.value = formula if not cell.value == 'AVG' else "AVG"
                for i, cell in enumerate(ws['E'], start=1):  # vice versa
                    formula = f'=IFERROR(STDEV(H{i}:{last_col_literal}{i}), "N/A")'
                    cell.value = formula if not cell.value == 'STD' else "STD"
                for i, cell in enumerate(ws['F'], start=1):
                    formula = f'=IFERROR(MIN((C{i} - D{i}) / (3 * E{i}),(D{i} - B{i})/ (3 * E{i})),"N/A")'
                    cell.value = formula if not cell.value == 'CPK' else "CPK"
                red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                for cell in ws['G']:
                    cell_letter = cell.coordinate
                    if cell_letter != 'G1' and float(cell.value) > 30:
                        ws.conditional_formatting.add(cell_letter, CellIsRule(operator='greaterThan', formula=['30'], fill=red_fill))

            wb.save(f'NPI_GRR_Result_{formatted_time}.xlsx')
            # workbook = xls.book
            # df.to_excel(xls, sheet_name='GRR')  # change sheet name
            # for s_name in range(2, 6):
            #     new_excel_filename = f'GRR{s_name}'
            #     # Copy all values from the source worksheet to the destination worksheet
            #     workbook_clone = workbook.add_worksheet(new_excel_filename)
            #     df_GRR = self.grr_weigh_pack(df, s_name)
            #     self.xlsxInstance.clone_by_weigh(workbook_clone=workbook_clone, shape=df.shape, sheet_namaiwa=f'GRR{s_name}', new_excel_filename=new_excel_filename, weigh=s_name)
            #     for r, row in enumerate(df_GRR.iterrows(), start=1):
            #         for c, value in enumerate(row[1], start=1):
            #             if isinstance(value, (int, float)) and (pd.isnull(value) or np.isinf(value)):
            #                 value = None  # Replace NaN/Inf with None
            #             else:
            #                 workbook_clone.write(r, c, value)
            #
            # xls.close()
            # wb.save(result_file_path)
        except Exception as e:
            raise "xlsx file NG >>>" + str(e.args)

    def grr_weigh_pack(self, *args):
        # pure numeric handling
        grr_lst = []
        try:
            data, avg_weigh = args
            result_arr = []
            df = data[~data.applymap(lambda x: isinstance(x, str)).any(axis=1)]
            for i, row in df.iterrows():
                extend_arr = row[6: df.shape[1]]
                if any(extend_arr.apply(lambda x: isinstance(x, str))):
                    grr_lst.append(np.nan)
                    df.drop(i, inplace=True)
                    continue
                # accumulate collection to extend arr for moving average
                extend_arr.loc[df.shape[1] - 6] = extend_arr[0]  # 6 => usl, lsl, etc cols
                if avg_weigh > 2:
                    extend_arr.loc[df.shape[1] - 6 + 1] = extend_arr[1]
                if avg_weigh > 3:
                    extend_arr.loc[df.shape[1] - 6 + 2] = extend_arr[2]
                if avg_weigh > 4:
                    extend_arr.loc[df.shape[1] - 6 + 3] = extend_arr[3]
                select_arr_ith = extend_arr.rolling(avg_weigh, min_periods=avg_weigh, center=False).mean().values
                select_arr_ith = select_arr_ith[avg_weigh - 1:]  # select non nan part made from rolling()
                result_arr.append(select_arr_ith)
                grr_arr = np.array(select_arr_ith[:90]).reshape(10, 9)
                grr_lst.append(self.grr_calculation(grr_arr, df['USL'][i], df['LSL'][i]))

            df_weigh = pd.DataFrame(result_arr)
            # if hasLiterals:
            #     df.drop(removeItem, inplace=True)
            df_weigh.index = df.index
            df_weigh.insert(loc=0, column='LSL', value=df['LSL'])
            df_weigh.insert(loc=1, column='USL', value=df['USL'])
            df_weigh.insert(loc=2, column='AVG', value="")
            df_weigh.insert(loc=3, column='STD', value="")
            df_weigh.insert(loc=4, column='CPK', value="")
            df_weigh.insert(loc=5, column=f'GRR{avg_weigh}', value=np.array(grr_lst))
        except Exception as e:
            raise "grr_weigh_pack$NG >>> " + str(e.args)
        return df_weigh

    def grr_summary(self, *args):
        arr_result = []
        path = ""
        current_time = datetime.datetime.now()
        formatted_time = current_time.strftime(f'%H%M%S')
        try:
            grr_file_path = args
            grr_files = os.listdir(grr_file_path[0])
            for file in grr_files:
                if file.endswith(".csv"):
                    path = os.path.join(grr_file_path[0], file)
                df = pd.read_csv(path, header=0, index_col=0)
                df_select = df.iloc[2:4, 1:]
                df_select = df_select.T
                df_select.columns = df_select.iloc[0]
                df_select.drop(df_select.index[0])
                arr_result.append(df_select)

            df_result = pd.concat([pd.DataFrame(arr) for arr in arr_result], axis=1)
            df_result = df_result.sort_index(axis=1, ascending=False)  # reverse sorting
            _filename = os.path.join(grr_file_path[0], f'Summary_GRR_{formatted_time}.csv')
            df_result.to_csv(_filename, sep=',')

        except Exception as e:
            raise "grr_summary NG >>> " + str(e.args)

    def grr_selection(self, *args):
        avg_result_lst = []
        grr_result_lst = []
        # path = ""
        current_time = datetime.datetime.now()
        formatted_time = current_time.strftime(f'%H%M%S')
        df_result = pd.DataFrame()
        try:
            summary_file_path = args[0]
            grr_csv_gather = os.listdir(summary_file_path)
            for file in grr_csv_gather:
                if file.startswith("Summary"):
                    p = os.path.join(summary_file_path, file)
                    df_result = pd.read_csv(p, header=0, index_col=0)
            # origin_col_name = df_result.columns
            df_result.columns = [i for i in range(df_result.shape[1])]
            df_result = df_result.drop(df_result.index[0])  # delete first redundant row
            for index, row in df_result.iterrows():
                _selected_weigh = row.iloc[-1]
                grr_result_lst.append(row.iloc[int(_selected_weigh)])
                # avg_result_lst.append(row.iloc[int(_selected_weigh) + 4])
            df_result['Final_GRR'] = grr_result_lst
            # df_result['Final_AVG'] = avg_result_lst
            _filename = os.path.join(summary_file_path, f'GRR_Result_{formatted_time}.csv')
            # df_result.columns = origin_col_name
            df_result.to_csv(_filename, sep=',')
            # XlsxManager.highlight_NG(df_result)

        except Exception as e:
            raise "grr_selection NG >>> " + str(e.args)
        return ""

    def convertToDf(*args) -> pd.DataFrame():
        delimiters = ('\t', ',')
        output_rows = []
        record = []
        output_file = 'output.csv'
        lines = args[0].strip().split('\n')
        try:
            for line in lines:
                line = line.strip()
                if not line:
                    continue

                if line[0].isdigit():
                    if record:
                        output_rows.append(record)

                    record = []

                elements = [line]
                for delimiter in delimiters:
                    if delimiter in line:
                        elements = line.split(delimiter)
                        break

                # Add the elements to record
                record.extend(elements)
            if record:
                output_rows.append(record)

            df = pd.DataFrame(output_rows)
            return df
        except Exception as ee:
            raise "convertToDf Err" + str(ee.args)

    # region    Corresponding Tx, Rx, Beam test items
    """
    Tx          --->    ( 7 ~ 14 + 36 ~ 65 ) makeTxTable
    Rx          --->    ( 15 ~ 26 + 66 ~ 105 ) makeRxTable
    BeamForm    --->    ( 105 ~ 109 ) makeBeamFormTable
    """

    def makeTxTable(sourcePath):
        filepath = os.path.join(sourcePath, f'RawData.csv')
        df = pd.read_csv(filepath, header=0, sep=',', encoding='UTF-8')
        subset1 = df.loc[7:14]
        subset2 = df.loc[34:65]
        Tx_rows = pd.concat([subset1, subset2])
        # Tx_rows = df.loc[7:14].append(df.loc[34:65])
        Tx_rows = Tx_rows.reset_index(drop=True)
        columnNames = {"0": 'TestItem', "1": 'Frequency', "2": 'DataRate', "3": 'Bandwidth', "4": 'Antenna',
                       "5": 'TxPower',
                       "6": 'Power',
                       "7": 'MaskMargin', "8": 'Freq1', "9": 'Freq2', "10": 'Freq3', "11": 'Freq4', "12": 'Freq5',
                       "13": 'Freq6',
                       "14": 'Freq7', "15": 'Freq8', "16": 'EVM', "17": 'FreqErr', "18": 'SpectrumMask',
                       "19": 'TestTime'}
        Tx_rows = Tx_rows.rename(columns=columnNames)
        TxDf = Tx_rows.iloc[:, 0:21]
        # TxTable = pd.DataFrame()
        # TxDf['TestItem'] = TxDf['TestItem']
        TxDf['Frequency'] = TxDf['Frequency'].str.split("Frequency: ").str[1]
        TxDf['DataRate'] = TxDf['DataRate'].str.split("Data Rate: ").str[1]
        TxDf['Bandwidth'] = TxDf['Bandwidth'].str.split("Bandwidth: BW-").str[1]
        TxDf['Antenna'] = TxDf['Antenna'].str.split("Antenna: ANT_").str[1]
        TxDf['TxPower'] = TxDf['TxPower'].str.split("Tx Power: ").str[1]
        TxDf['Power'] = TxDf['Power'].str.split("Power             ").str[1]
        # TxDf['MaskMargin'] = TxDf['MaskMargin']
        TxDf['Freq1'] = TxDf['Freq1'].str.split("Frequency          ").str[1]
        TxDf['Freq2'] = TxDf['Freq2'].str.split("Frequency          ").str[1]
        TxDf['Freq3'] = TxDf['Freq3'].str.split("Frequency          ").str[1]
        TxDf['Freq4'] = TxDf['Freq4'].str.split("Frequency          ").str[1]
        TxDf['Freq5'] = TxDf['Freq5'].str.split("Frequency          ").str[1]
        TxDf['Freq6'] = TxDf['Freq6'].str.split("Frequency          ").str[1]
        TxDf['Freq7'] = TxDf['Freq7'].str.split("Frequency          ").str[1]
        TxDf['Freq8'] = TxDf['Freq8'].str.split("Frequency          ").str[1]
        TxDf['EVM'] = TxDf['EVM'].str.split("EVM          ").str[1]
        TxDf['FreqErr'] = TxDf['FreqErr'].str.split("Freq Error          ").str[1]
        TxDf['SpectrumMask'] = TxDf['SpectrumMask'].str.split("Spectrum Mask   ").str[1]
        TxDf['TestTime'] = TxDf['TestTime'].str.split("Test time: ").str[1]
        # TxDf.head()
        csvpath = os.path.join(sourcePath, "Data")
        TxDf.to_csv(os.path.join(csvpath, f'TxDF.csv'), sep=',', encoding='UTF-8')

    def makeRxTable(sourcePath):
        filepath = os.path.join(sourcePath, f'RawData.csv')
        df = pd.read_csv(filepath, header=0, sep=',', encoding='UTF-8')
        subset1 = df.loc[15:26]
        subset2 = df.loc[66:105]
        Rx_rows = pd.concat([subset1, subset2])
        Rx_rows = Rx_rows.reset_index(drop=True)
        columnNames = {"0": 'TestItem', "1": 'Frequency', "2": 'DataRate', "3": 'Bandwidth',
                       "4": 'Antenna', "5": 'RxPower', "6": 'Per', "7": 'TestTime'}

        RxDf = Rx_rows.iloc[:, 0:9]
        RxDf = RxDf.rename(columns=columnNames, inplace=False)
        # RxTable = pd.DataFrame()
        # RxDf['TestItem'] = RxDf['TestItem']
        RxDf['Frequency'] = RxDf['Frequency'].str.split("Frequency: ").str[1]
        RxDf['DataRate'] = RxDf['DataRate'].str.split("Data Rate: ").str[1]
        RxDf['Bandwidth'] = RxDf['Bandwidth'].str.split("Bandwidth: BW-").str[1]
        RxDf['Antenna'] = RxDf['Antenna'].str.split("Antenna: ANT_").str[1]
        RxDf['RxPower'] = RxDf['RxPower'].str.split("Rx Power: ").str[1]
        RxDf['Per'] = RxDf['Per'].str.split("PER    ").str[1]
        RxDf['PER_result'] = RxDf['Per'].str.split("<-- ").str[1]
        RxDf['TestTime'] = RxDf['TestTime'].str.split("Test time: ").str[1]
        csvpath = os.path.join(sourcePath, "Data")
        RxDf.to_csv(os.path.join(csvpath, f'RxDF.csv'), sep=',', encoding='UTF-8')

    def makeBeamFormTable(sourcePath):
        filepath = os.path.join(sourcePath, f'RawData.csv')
        df = pd.read_csv(filepath, header=0, sep=',', encoding='UTF-8')
        WifiBeamForming_rows = df.loc[106:109]
        WifiBeamForming_rows = WifiBeamForming_rows.reset_index(drop=True)
        # Create a dictionary to map the old column names to the new column names
        columnNames = {"0": 'TestItem', "1": 'TxFrequency', "2": 'TxDataRate', "3": 'TxBandwidth', "4": 'TxAntenna',
                       "5": 'TxPower', "6": 'RxFrequency', "7": 'RxDataRate', "8": 'RxBandwidth', "9": 'RxAntenna',
                       "10": 'RxPower', "11": 'rPcal', "12": 'rPcal_core3', "13": 'TxBF', "14": 'RxFreq',
                       "15": 'RxDR', "16": 'RxBw', "17": 'RxAnt', "18": 'RxPow', "19": 'TxFreq',
                       "20": 'TxDR', "21": 'TxBw', "22": 'TxAnt', "23": 'TxPow', "24": 'PowerDiff', "25": 'TestTime'}

        WifiBeamForming_rows = WifiBeamForming_rows.rename(columns=columnNames)
        WifiBeamForming_rows = WifiBeamForming_rows.iloc[:, 0:27]
        WifiBeamForming_Table = pd.DataFrame()
        WifiBeamForming_Table['TestItem'] = WifiBeamForming_rows['TestItem']
        WifiBeamForming_Table['TxFrequency'] = WifiBeamForming_rows['TxFrequency'].str.split("Tx Frequency: ").str[1]
        WifiBeamForming_Table['TxDataRate'] = WifiBeamForming_rows['TxDataRate'].str.split("Data Rate: ").str[1]
        WifiBeamForming_Table['TxBandwidth'] = WifiBeamForming_rows['TxBandwidth'].str.split("Bandwidth: ").str[1]
        WifiBeamForming_Table['TxAntenna'] = WifiBeamForming_rows['TxAntenna'].str.split("Antenna: ").str[1]
        WifiBeamForming_Table['TxPower'] = WifiBeamForming_rows['TxPower'].str.split("Tx Power: ").str[1]
        WifiBeamForming_Table['RxFrequency'] = WifiBeamForming_rows['RxFrequency'].str.split("Rx Frequency: ").str[1]
        WifiBeamForming_Table['RxDataRate'] = WifiBeamForming_rows['RxDataRate'].str.split("Data Rate: ").str[1]
        WifiBeamForming_Table['RxBandwidth'] = WifiBeamForming_rows['RxBandwidth'].str.split("Bandwidth: ").str[1]
        WifiBeamForming_Table['RxAntenna'] = WifiBeamForming_rows['RxAntenna'].str.split("Antenna: ").str[1]
        WifiBeamForming_Table['RxPower'] = WifiBeamForming_rows['RxPower'].str.split("Rx Power: ").str[1]
        WifiBeamForming_Table['rPcal'] = WifiBeamForming_rows['rPcal'].str.split("rPcal: ").str[1]
        WifiBeamForming_Table['rPcal_core3'] = WifiBeamForming_rows['rPcal_core3'].str.split("rPcal_core3: ").str[1]
        WifiBeamForming_Table['TxBF'] = WifiBeamForming_rows['TxBF'].str.split("TxBF Cal            ").str[1]
        WifiBeamForming_Table['RxFreq'] = WifiBeamForming_rows['RxFreq'].str.split("Rx Frequency: ").str[1]
        WifiBeamForming_Table['RxDR'] = WifiBeamForming_rows['RxDR'].str.split("Data Rate: ").str[1]
        WifiBeamForming_Table['RxBw'] = WifiBeamForming_rows['RxBw'].str.split("Bandwidth: ").str[1]
        WifiBeamForming_Table['RxAnt'] = WifiBeamForming_rows['RxAnt'].str.split("Antenna: ").str[1]
        WifiBeamForming_Table['RxPow'] = WifiBeamForming_rows['RxPow'].str.split("Rx Power: ").str[1]
        WifiBeamForming_Table['TxFreq'] = WifiBeamForming_rows['TxFreq'].str.split("Tx Frequency: ").str[1]
        WifiBeamForming_Table['TxDR'] = WifiBeamForming_rows['TxDR'].str.split("Data Rate: ").str[1]
        WifiBeamForming_Table['TxBw'] = WifiBeamForming_rows['TxBw'].str.split("Bandwidth: ").str[1]
        WifiBeamForming_Table['TxAnt'] = WifiBeamForming_rows['TxAnt'].str.split("Antenna: ").str[1]
        WifiBeamForming_Table['TxPow'] = WifiBeamForming_rows['TxPow'].str.split("Tx Power: ").str[1]
        WifiBeamForming_Table['PowerDiff'] = WifiBeamForming_rows['PowerDiff'].str.split("Power Diff          ").str[1]
        WifiBeamForming_Table['TestTime'] = WifiBeamForming_rows['TestTime'].str.split("Test time: ").str[1]
        # WifiBeamForming_Table.head()
        csvpath = os.path.join(sourcePath, "Data")
        WifiBeamForming_Table.to_csv(os.path.join(csvpath, f'BeamForming_Table.csv'), sep=',', encoding='UTF-8')

    # endregion

    # region    Corresponding Tx, Rx CALIBRATION items
    """
    Corresponding test items
    {
        5. WIFI_RX_CALIBRATION  ---> makeWifiRxCalib5   ,  
        6. WIFI_TX_CALIBRATION  ---> makeWifiTxCalib6   ,
        32. WIFI_RX_CALIBRATION ---> makeWifiRxCalib32  ,
        33. WIFI_TX_CALIBRATION ---> makeWifiTxCalib33
    }
    """

    def makeWifiTxCalib33(sourcePath):
        """
        # The testItem of 33. WIFI_TX_CALIBRATION as a CSV which contain two tables
          * Calibration TX Power
          * Verify TX Power
        """
        excel_file = 'WIFI_TX_CALIBRATION33.xlsx'
        dfsCalfTXPow = pd.DataFrame()
        dfsVerfTXPow = pd.DataFrame()
        filepath = os.path.join(sourcePath, f'RawData.csv')
        xlsxPath = os.path.join(os.path.join(sourcePath, "Data"), f'{excel_file}')
        try:
            df = pd.read_csv(filepath, header=0, sep=',', encoding='UTF-8')
            row33 = df.loc[33].transpose().dropna().reset_index(drop=True)
            # row33 = row33[not row33.str.contains("-----------")].reset_index(drop=True)
            CalibrationTXPower = row33.loc[5:72]
            VerifyTXPower = row33.loc[76:144]
            dfsCalfTXPow[['Rate', 'Ant', 'Target', 'Actual', 'Diff']] = CalibrationTXPower.iloc[:].str.split('\s+',
                                                                                                             expand=True)
            dfsVerfTXPow[['Rate', 'Ant', 'Target', 'Actual']] = VerifyTXPower.iloc[:].str.split('\s+', expand=True)
            dfsCalfTXPow.reset_index(drop=True)
            dfsVerfTXPow.reset_index(drop=True)
            dfsCalfTXPow.loc[25:, ['Actual', 'Diff']] = dfsCalfTXPow.loc[25:, ['Diff', 'Actual']].values
            with pd.ExcelWriter(xlsxPath, engine='openpyxl') as writer:
                dfsCalfTXPow.to_excel(writer, sheet_name='CalibrationTXPower', index=False)
                dfsVerfTXPow.to_excel(writer, sheet_name='VerifyTXPower', index=False)
        except SyntaxError as synEx:
            raise "err! " + str(synEx.args)
        except Exception as ee:
            raise f"make Tx Calif excel in valid\n" + str(ee.args)

    def makeWifiTxCalib6(sourcePath):
        """
        # The testItem of 6. WIFI_TX_CALIBRATION as a xlsx which contain two excel table sheets
          * Calibration TX Power
          * Verify TX Power
        """
        excel_file = 'WIFI_TX_CALIBRATION6.xlsx'
        dfs6Calf = pd.DataFrame()
        dfs6Verf = pd.DataFrame()
        filepath = os.path.join(sourcePath, f'RawData.csv')
        xlsxPath = os.path.join(os.path.join(sourcePath, "Data"), f'{excel_file}')
        try:
            df = pd.read_csv(filepath, header=0, sep=',', encoding='UTF-8')
            row6 = df.loc[6].transpose().dropna().reset_index(drop=True)
            # row6 = row6[~row6.str.contains("-----------")].reset_index(drop=True)
            CalibTXPower = row6.loc[5:16]
            VerifyTXPower = row6.loc[20:31]
            dfs6Calf[['Rate', 'Ant', 'Target', 'Actual', 'Diff']] = CalibTXPower.iloc[:].str.split('\s+', expand=True)
            dfs6Verf[['Rate', 'Ant', 'Target', 'Actual']] = VerifyTXPower.iloc[:].str.split('\s+', expand=True)
            dfs6Calf.reset_index(drop=True)
            dfs6Verf.reset_index(drop=True)
            dfs6Calf.loc[9:, ['Actual', 'Diff']] = dfs6Calf.loc[9:, ['Diff', 'Actual']].values
            with pd.ExcelWriter(xlsxPath, engine='openpyxl') as writer:
                dfs6Calf.to_excel(writer, sheet_name='CalibrationTXPower', index=False)
                dfs6Verf.to_excel(writer, sheet_name='VerifyTXPower', index=False)

        except SyntaxError as synEx:
            raise "err! " + str(synEx.args)
        except Exception as ee:
            raise f"make Tx Calif excel in valid\n" + str(ee.args)

    def makeWifiRxCalib5(sourcePath):
        """
        # 5. WIFI_RX_CALIBRATION as a csv
          * Calibration TX Power
          * Verify TX Power
        """
        filepath = os.path.join(sourcePath, f'RawData.csv')
        try:
            df = pd.read_csv(filepath, header=0, sep=',', encoding='UTF-8')
            WIFI_RX_CALIBRATION = df.loc[5].transpose().dropna().reset_index(drop=True)
            # selectHeader = WIFI_RX_CALIBRATION.loc[0:1].reset_index(drop=True)
            lastIndex = WIFI_RX_CALIBRATION.index[-1]
            selectFront = WIFI_RX_CALIBRATION.loc[2:5]
            selectBack = WIFI_RX_CALIBRATION.loc[7:10]
            selectTestTime = WIFI_RX_CALIBRATION.loc[lastIndex:]
            itemsFront = selectFront.iloc[:].str.split(":", expand=True)
            itemsFront[1] = itemsFront[1].apply(lambda s: s.split())
            itemsBack = selectBack.iloc[:].str.split(":", expand=True).reset_index(drop=True)
            itemsTestTime = selectTestTime.iloc[:].str.split(":", expand=True).reset_index(drop=True)
            itemsBack.loc[1:, [1, 0]] = itemsBack.loc[1:, [0, 1]].values
            itemsBack[1].loc[0] = itemsBack[1].agg(lambda x: x.tolist())
            itemsBack = itemsBack.loc[[0]].reset_index(drop=True)
            result = pd.concat([itemsFront, itemsBack, itemsTestTime])  # selectHeader
            result.columns = ['keys', 'values']
            result.reset_index(drop=True)
            # print(result)
            csvpath = os.path.join(sourcePath, "Data")
            result.to_csv(os.path.join(csvpath, f'WIFI_RX_CALIBRATION_5.csv'), sep=',', encoding='UTF-8')
        except Exception as ee:
            raise f"make Tx Calif5 excel in valid\n" + str(ee.args)

    def selectRow(rowRange: list, filterData, isSplit: bool):
        selected_rows = pd.Series(False, index=filterData.index)
        result = None
        for start, end in rowRange:
            selected_rows[start:end] = True
        selected = filterData[selected_rows]
        if isSplit:
            rowValues = selected.iloc[:].str.split(":", expand=True)
            rowValues[1] = rowValues[1].apply(lambda s: s.split())
            print(rowValues)
            return rowValues
        else:
            result = selected.iloc[:].str.split(":", expand=True).reset_index(drop=True)
            row_ranges = [(1, 3), (5, 7), (9, 11)]
            last_row_range = (13, None)
            # exchanges ------------
            for start, end in row_ranges:
                result.loc[start:end, [1, 0]] = result.loc[start:end, [0, 1]].values
            result.loc[last_row_range[0]:, [1, 0]] = result.loc[last_row_range[0]:, [0, 1]].values
            # exchanges ------------
            row_ranges = [(0, 3), (4, 7), (8, 11)]
            last_row_range = (12, None)
            # melting to combine
            for i, (start, end) in enumerate(row_ranges):
                result[1].loc[i * 4] = result[1].loc[start:end].agg(lambda x: x.tolist())

            result[1].loc[12] = result[1].loc[last_row_range[0]:].agg(lambda x: x.tolist())
            result = result.loc[[0, 4, 8, 12]].reset_index(drop=True)
            result = result.reset_index(drop=True)
            return result

    def makeWifiRxCalib32(sourcePath):
        """
        # 32. WIFI_RX_CALIBRATION as a csv
          * Calibration TX Power
          * Verify TX Power
        """
        filepath = os.path.join(sourcePath, f'RawData.csv')
        try:
            df = pd.read_csv(filepath, header=0, sep=',', encoding='UTF-8')
            WIFI_RX_CALIBRATION = df.loc[32].transpose().dropna().reset_index(drop=True)
            data = WIFI_RX_CALIBRATION.loc[1:52]
            filterData = data[~data.str.contains("-----------|========")].reset_index(drop=True)
            row_ranges = [(1, 7), (12, 18), (23, 26), (31, 37)]
            row_FinalGainErr = [(8, 12), (19, 23), (27, 31), (38, 42)]
            lastIndex = WIFI_RX_CALIBRATION.index[-1]
            selectTestTime = WIFI_RX_CALIBRATION.loc[lastIndex:]  # .loc[[-1]].reset_index(drop=True)
            itemsFront = Digest_utils.selectRow(rowRange=row_ranges, filterData=filterData, isSplit=True)
            itemsBack = Digest_utils.selectRow(rowRange=row_FinalGainErr, filterData=filterData, isSplit=False)

            itemsTestTime = selectTestTime.iloc[:].str.split(":", expand=True).reset_index(drop=True)
            result = pd.concat([itemsFront, itemsBack, itemsTestTime])  # selectHeader
            result.columns = ['keys', 'values']
            csvpath = os.path.join(sourcePath, "Data")
            result.reset_index(drop=True)
            result.to_csv(os.path.join(csvpath, f'WIFI_RX_CALIBRATION_32.csv'), sep=',', encoding='UTF-8')
        except Exception as ee:
            raise f"make Tx Calif32 excel in valid\n" + str(ee.args)

    def digestFiles(self, file_path, csv_writer, file_name):
        string_to_find = "SPEC:"
        string_to_find2 = "Value:"
        try:
            with open(file_path, 'r', encoding='utf-16') as file:
                # Iterate through lines in the file
                for line_number, line in enumerate(file, start=1):
                    # Check if the line contains the specified string
                    if string_to_find in line and string_to_find2 in line:
                        # Write the file name and line to the CSV file
                        csv_writer.writerow([file_name, f'{line.strip()}'])
        except UnicodeDecodeError:
            print(f"Error decoding file {file_name}. Skipping...")
        except Exception as e:
            raise "digestFiles$NG >>> " + str(e.args)

    def NG_Open_log_txt(self, directory_path, output_csv_path):
        keyword1 = "SPEC:"
        keyword2 = "Value:"
        ending_keyword = "END MARKED"
        USL_lst = []
        LSL_lst = []
        testItem_lst = []
        pattern = r"~"
        try:
            with open(output_csv_path, 'w', newline='', encoding='utf-16') as txtFile:
                # csv_writer = csv.writer(csvfile)
                # Iterate through files in the specified directory
                for filename in os.listdir(directory_path):
                    file_path = os.path.join(directory_path, filename)
                if os.path.isfile(file_path):
                    with open(file_path, 'r', encoding='utf-16') as file:
                        for line_number, line in enumerate(file, start=1):
                            if keyword1 in line and keyword2 in line:
                                txtFile.write(f'{line.strip()}\n')
                                # Write the file name and line to the CSV file
                                # csv_writer.writerow([file_name, f'{line.strip()}'])
                            if ending_keyword in line:  # stop reading til the end keyword
                                break
                            # Find the indices of the keywords in the line
                            index1 = line.find(keyword1)
                            index2 = line.find(keyword2)
                            test_item_name = line[:index1].strip(' ').split(' ')[-1]
                            testItem_lst.append(test_item_name)
                            specRange = line[index1 + len(keyword1):index2].strip()
                            if not re.search(pattern, specRange):
                                USL_lst.append("PASS")
                                LSL_lst.append("PASS")
                            else:
                                LSL_lst.append(specRange.split('~')[0])
                                USL_lst.append(specRange.split('~')[1])
                                # Write the file name and line to the CSV file
                                # csv_writer.writerow([file_name, f'{line.strip()}']
            # file.close()
            txtFile.close()
            return LSL_lst, USL_lst, testItem_lst
        except UnicodeDecodeError:
            print(f"Error decoding file {txtFile}. Skipping...")
            raise "UnicodeDecodeError"
        except io.UnsupportedOperation:
            print("not readable")
            raise "io exception"
        except Exception as e:
            raise "digestFiles$NG >>> " + str(e.args)

    def Open_log_txt(self, file_path, output_csv_path):
        global testItem_lst
        # global USL_lst
        # global LSL_lst
        keyword1 = "SPEC:"
        keyword2 = "Value:"
        flag = False
        USL_lst = []
        LSL_lst = []
        testItem_lst = []
        countStressTimes = 0
        pattern = r"~"
        ending_keyword = "END MARKED"
        proj = "IMQX"
        try:
            with open(output_csv_path, 'w', newline='', encoding='utf-16') as txtFile:
                # csv_writer = csv.writer(csvfile)
                if os.path.isfile(file_path):
                    with open(file_path, 'r', encoding='utf-16') as file:
                        for line_number, line in enumerate(file, start=1):
                            if ending_keyword in line:  # stop reading til the end keyword
                                countStressTimes += 1
                                flag = True
                            if keyword1 in line and keyword2 in line:
                                # Find the indices of the keywords in the line
                                index1 = line.find(keyword1)
                                index2 = line.find(keyword2)
                                if not flag:
                                    test_item_name = line[:index1].strip(' ').split(' ')[-1]
                                    testItem_lst.append(test_item_name)
                                    specRange = line[index1 + len(keyword1):index2].strip()
                                    if not re.search(pattern, specRange):
                                        USL_lst.append("PASS")
                                        LSL_lst.append("PASS")
                                    else:
                                        LSL_lst.append(specRange.split('~')[0])
                                        USL_lst.append(specRange.split('~')[1])
                                txtFile.write(f'{line.strip()}\n')
                            # Write the file name and line to the CSV file
                            # csv_writer.writerow([file_name, f'{line.strip()}'])
        except IOError as ioe:
            print(f'Open_log_txt$IOError >>> {ioe.args}')
        except UnicodeDecodeError as ue:
            print(f"Open_log_txt$decoding file {file} >>>> {ue.args} \r\n  Skipping...")
        except Exception as e:
            raise "digestFiles$NG >>> " + str(e.args)
        print("----------------")
        result = (countStressTimes, len(testItem_lst), LSL_lst, USL_lst)
        return result

    def washing(self, file_path, df_info):
        """
        Diagnosis >>> count test items
        :return:
        """
        global testItem_lst
        keyword1 = "SPEC:"
        keyword2 = "Value:"
        all_lst = []
        LSL_lst = df_info[2]
        USL_lst = df_info[3]
        data_container = {item: None for item in testItem_lst}
        num_lost_lines = 0
        num_NG = 0
        isNG_Occur = False
        heap_testItem_lst = []
        stack_testItem_lst = testItem_lst.copy()
        try:
            with open(file_path, mode='r', encoding='utf-16', errors='ignore') as file:
                lines = file.readlines()
                for ith, line in enumerate(lines, start=1):
                    if "TestWarning" in line or "EngMode" in line or "RetryTimes" in line:
                        print("engineering mode, invalid parsing")
                        continue
                    # Find the indices of the keywords in the line
                    index1 = line.find(keyword1)
                    index2 = line.find(keyword2)
                    # num_lost_lines += num_lost_lines
                    # TestItem_nth = (ith + num_lost_lines) % countTestItems
                    test_item_name = line[:index1].strip(' ').split(' ')[-1]
                    testItem_values = line[index2 + len(keyword2):].strip()
                    # num_lost_lines = testItem_lst.index(test_item_name)
                    # # data_lst = data_lst + [np.nan] * abs(num_lost_lines)
                    # stack_testItem_lst.remove(test_item_name)
                    if test_item_name not in testItem_lst:
                        print("--- Ignore singular testItem ---")
                        num_lost_lines -= 1
                        continue
                    elif isNG_Occur and test_item_name == testItem_lst[0]:
                        print("--- jump next round after encounter NG ---")
                        num_NG += 1
                        stack_testItem_lst.clear()
                    elif test_item_name in heap_testItem_lst:
                        print("--- involving retry ---")
                        data_container[test_item_name] = testItem_values
                    else:
                        item = stack_testItem_lst.pop(0)
                        item_lsl = LSL_lst[testItem_lst.index(f'{item}')]
                        item_usl = USL_lst[testItem_lst.index(f'{item}')]
                        heap_testItem_lst.append(item)  # testItem already done
                        data_container[test_item_name] = testItem_values

                        if item_lsl == item_usl:
                            print("---- USL = LSL ----")
                        elif self.Instance_try_parse(item_usl) or self.Instance_try_parse(item_lsl) or self.Instance_try_parse(testItem_values):
                            print(" ---- none-sense spec or spec => literals---- ")
                        elif float(testItem_values) > float(item_usl) or float(testItem_values) < float(item_lsl):
                            isNG_Occur = True
                    if len(stack_testItem_lst) == 0:  # len(data_lst) == len(testItem_lst):
                        # if all(np.isnan(x) for x in np.array(data_lst).astype(float) if isinstance(x, (int, float))):
                        #     raise Exception("all values NAN")
                        # all_lst.append(data_lst)  # all_array = np.concatenate((all_array, np.array([data_lst])), axis=0)
                        all_lst.append(data_container)
                        heap_testItem_lst.clear()
                        stack_testItem_lst = testItem_lst.copy()
                        data_container = {item: None for item in stack_testItem_lst}
                # df = self.pack_result(all_lst, LSL_lst, USL_lst, result_fileName)
                return all_lst, LSL_lst, USL_lst  # df, num_NG
        except Exception as e:
            raise "washing$NG >>> " + str(e.args)
        finally:
            file.close()

    @staticmethod
    def Instance_try_parse(literal) -> bool:
        try:
            # isinstance(item_usl, str) or isinstance(item_usl, str)
            float(literal)
            return False
        except Exception as e:
            print("pure string" + str(e.args))
            return True

    def pack_result(self, result_list, LSL_lst, USL_lst, result_fileName) -> pd.DataFrame():

        try:
            # df = pd.DataFrame(np.array(result_list))
            # df.columns = pd.Series(testItem_lst)
            df = pd.DataFrame.from_dict(result_list)
            df = df.T
            df = df.dropna(axis=1, how='all')
            df = df.dropna(axis=0, how='all')
            df.insert(loc=0, column='LSL', value=LSL_lst)
            df.insert(loc=1, column='USL', value=USL_lst)
            df.insert(loc=2, column='AVG', value="")
            df.insert(loc=3, column='STD', value="")
            df.insert(loc=4, column='CPK', value="")
            result_cpk_xlsx = os.path.join(self.output_path, result_fileName)
            writer = pd.ExcelWriter(f'{result_cpk_xlsx}.xlsx', engine='xlsxwriter', engine_kwargs={'options': {'strings_to_numbers': True}})
            df.to_excel(writer, sheet_name='GRR')
            self.xlsxInstance.cooking_xCPK(writer=writer, shape=df.shape)
            return df
        except IOError as e:
            print("Xlsx File already open! >>> " + str(e.args))
        except Exception as e:
            raise "pack_result$NG >>> " + str(e.args)

    def result_display(self, *args):
        try:
            switcher = {
                0: "Horizontal",
                1: "Vertical",
            }
            return switcher.get(args[0], "$$$")
        except Exception as e:
            raise "xlsx result display$NG >>>" + str(e.args)
        finally:
            print(args)

    def AcquireLatestFile(self):
        try:
            files = os.listdir(self.output_path)
            # Sort the files based on modification time (newest first)
            files.sort(key=lambda x: os.path.getmtime(os.path.join(self.output_path, x)), reverse=True)
            xlsx_path = os.path.join(self.output_path, files[0])
        except Exception as e:
            raise "AcquireLatestFile$NG >>>" + str(e.args)
        return xlsx_path
