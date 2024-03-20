from abc import ABCMeta, abstractmethod
from datetime import datetime
import openpyxl
import pandas as pd
import numpy as np
import json
import csv
import os


class DataOrder(metaclass=ABCMeta):

    def __init__(self):
        # self._chef = DataChef()
        self.__wifiType = None
        self.__csvPath = os.path.dirname(os.path.abspath(__file__))
        self.__csvfile = None
        self.__IsCsvExist = None
        self._df = pd.DataFrame()

    def __repr__(self):
        return f'__IsCsvExist {self.__IsCsvExist}'

    @abstractmethod
    def manifestData(self):
        """manifest Data column unique features"""
        pass

    @abstractmethod
    def processingCSV(self):
        pass

    @abstractmethod
    def queryResult(self):
        pass


class DataUtils(DataOrder):
    def queryResult(self):
        pass

    def manifestData(self, *args):
        # print(self.__doc__)
        pass

    @staticmethod
    def DigestPolishData(filePath: str, header) -> pd.DataFrame():
        try:
            if filePath is not None:
                # file_create_time = datetime.fromtimestamp(os.path.getctime(filePath)).strftime("%m%d_%H%M%S")
                base_name, extension = os.path.splitext(filePath)
                if extension == ".csv":
                    df = pd.read_csv(filePath, skiprows=range(1, 16), header=header, sep=',', encoding='UTF-8')
                elif extension == ".xlsx":
                    wb = openpyxl.load_workbook(filePath)
                    sheet_names = wb.sheetnames[0]
                    df = pd.read_excel(filePath, skiprows=range(1, 16), header=header, sheet_name=f'{sheet_names}')

            # df = pd.read_csv(filePath, header=0, sep=',', encoding='UTF-8',
            #                  usecols=['Result', 'Dut_No', 'type', 'channel', 'antenna', 'bandwidth',
            #                           'data_rate', 'EVM_Spec', 'EVM', 'Power_Spec', 'Power'])
        except Exception as ee:
            raise "DigestPolishData err" + str(ee.args)

        return df

    @staticmethod
    def DigestSketchData(filePath: str) -> pd.DataFrame():
        try:
            df = pd.read_csv(filePath, header=0, sep=',', encoding='UTF-8')
            # df = df[~(df['Result'] == 'TotalPass')].drop('Result', axis=1)
        except ValueError as Vee:
            raise "CSV format err" + str(Vee.args)
        except Exception as ee:
            raise "DigestSketchData err" + str(ee.args)

        return df

    def DigestAskOpenFileName(*args) -> pd.DataFrame():
        try:
            sourcePath = args[0]
            header = args[1]
            base_name, extension = os.path.splitext(sourcePath)
            if extension == ".csv":
                # csv_files = [file for file in os.listdir(filePath) if file.endswith(f'.csv')]
                # filePath = os.path.join(filePath, csv_files[-1])
                df = pd.read_csv(sourcePath, header=header, sep=',', encoding='UTF-8',
                                 skiprows=range(1, 16),
                                 usecols=['Result', 'Dut_No', 'type', 'channel', 'antenna', 'bandwidth', 'data_rate',
                                          'RX_PER'])
                df = df[~df['Result'].isin(['TotalPASS', 'TotalPass'])]
                # df = df[~(df['Result'] == 'TotalPASS')].drop('Result', axis=1)
                # df = df.drop(df[df['Result'] == 'TotalPass'].index, inplace=True)
            elif extension == ".xlsx":
                wb = openpyxl.load_workbook(sourcePath)
                sheet_names = wb.sheetnames[0]
                df = pd.read_excel(sourcePath, sheet_name=sheet_names, header=header, na_values=['NA'],
                                   skiprows=range(1, 16), usecols="B:H,N")
                df = df[~df['Result'].isin(['TotalPASS', 'TotalPass'])]
                # dfs = dfs.drop('Result', axis=1)
            else:
                raise f'{sourcePath} unacceptable'
            # ---------------------------------------------------------------
            # wifiCategory = df['type'].value_counts().index  # .sort_index()
            # Append newline to each index
            # wifiCategoryindex = wifiCategory.map(lambda x: str(x) + '\n')
            # new_counts = pd.Series(index=wifiCategory, data=list(wifiCategoryindex))
            # _counts = df.groupby('type')['Result'].value_counts().unstack()
        except Exception as RxErr:
            raise "DigestRx" + str(RxErr.args)
        return df

    @staticmethod
    def DigestRxTreemap(df: pd.DataFrame()):
        """
        :proportion -> Dut proportion
        :groups -> Strings in small square
        :countFail -> colorBar implication
        """
        groupLiteral = []
        df.reset_index(inplace=True, drop=True)
        try:
            proportion = df.groupby('Dut_No').size()
            fail_mask = df.groupby('Dut_No')['RX_PER'].apply(lambda x: (x > 0))
            divideNum = proportion.iloc[0]
            countFails = None
            for index in countFails.index:
                value = countFails.loc[index]
                groupLiteral.append(f"DutNo_{index}\n\n{str(value)}")
            groups = pd.Series(groupLiteral)
            groups.reset_index(drop=True)
            fail_indices = df.loc[fail_mask, :].index
            failIndexInDutNo = [int(index % (divideNum)) for index in fail_indices]
            dfs = pd.DataFrame({'proportion': proportion, 'group': groups, 'countFail': countFails})

        except Exception as treeR:
            raise "DigestTreemap" + str(treeR.args)
        return dfs

    @staticmethod
    def DigestTreemap(df: pd.DataFrame()):
        """
        :proportion -> Dut proportion
        :groups -> Strings in small square
        :countFail -> colorBar implication
        """
        groupLiteral = []
        df.reset_index(inplace=True, drop=True)
        try:
            proportion = df.groupby('Dut_No')['Result'].size()
            uniqueType = df.loc[df['Dut_No'] == 1, 'type']
            fail_mask = df.groupby('Dut_No')['Result'].apply(lambda x: (x == 'Fail'))
            divideNum = proportion.iloc[0]
            fail_indices = df.loc[fail_mask, :].index
            failIndexInDutNo = [int(index % (divideNum)) for index in fail_indices]

            # whichTestItemFail = failures.groupby('Dut_No').first()['type']
            # whichTestItemFail = proportion.sort_values().index
            # proportion = df.groupby(['Dut_No', 'Result']).size().unstack()
            # fail_proportions = proportion['Fail'] / proportion.sum(axis=1)
            # order = fail_proportions.sort_values().index

            # proportion = pd.Series(df['Dut_No'].value_counts().values)
            # countTypes = proportion.iloc[sorted(range(len(proportion)), key=lambda i: proportion.index[i])].values
            dut_no_fail = df.groupby('Dut_No')['Result'].value_counts().unstack().reindex(columns=['Fail', 'Pass'],
                                                                                          fill_value=0)
            if dut_no_fail['Fail'].isna().any():
                dut_no_fail['Fail'] = dut_no_fail['Fail'].fillna(0)
            countFails = dut_no_fail['Fail']
            countFails = countFails.apply(lambda x: failIndexInDutNo.pop() if x >= 1 else x)
            # countFail = df[df['Result'] == 'Fail'].groupby('Dut_No')['Result'].count()
            # countAll = df.loc[df['Result']].groupby('Dut_No')['Result'].value_counts()
            # failed_dut_nos = countFail[countFail > 0].index.tolist()
            fail_indices = df.groupby('Dut_No').apply(lambda x: x.index[x['Result'] == 'Fail'])
            fffff = df.loc[df['Result'] == 'Fail'].groupby('Dut_No')['Result'].rank(method='first')
            # df_fail = df[df['Result'] == 'Fail'].groupby('Dut_No').cumsum().sort_values(['Dut_No', 'Result'])
            # failed_dut_num = ["DutNo_" + str(item) + '\n' for item in failed_dut_nos]
            DutGroup = ["DutNo_" + str(item) + '\n' for item in df['Dut_No'].unique()]
            # [failed_dut_list.append(str(item) + '_suffix') for item in failed_dut_num]
            # failed_dut_nos = failed_dut_nos.apply(lambda x: "DutNo_" + str(x) + '\n')
            # groupName = countFail.apply(lambda x, : f"{x[0]} \n\n" + str(x[1]))

            for index in countFails.index:
                value = countFails.loc[index]
                groupLiteral.append(f"DutNo_{index}\n\n{str(value)}")
            groups = pd.Series(groupLiteral)
            # countFail_decomposed = pd.Series([f"DutNo_{key} \n\n {str(value)}" for key, value in countFail.items()])
            groups.reset_index(drop=True)
            groups.index = pd.RangeIndex(start=1, stop=len(groupLiteral) + 1)
            dfs = pd.DataFrame({'proportion': proportion, 'group': groups, 'countFail': countFails})
            # dfs['proportion'] = proportion
            # dfs['countFail'] = df[df['Result'] == 'Fail'].groupby('Dut_No')['Result'].count()

        except Exception as treeR:
            raise "DigestTreemap" + str(treeR.args)
        return dfs

    @staticmethod
    def fetchTSB():
        files = []
        root = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Data')
        for directory, subdirectories, files in os.walk(root):
            for file in files:
                files.append(os.path.join(directory, file))
                print(os.path.join(directory, file))
                return files

    def processingCsv(*args, **kwargs):
        WifiType = args[0]
        filePath = args[1]  # DirData
        PathForTxSketch = args[2]  # DirDataSketch
        PathForRxSketch = args[3]  # DirDataRx
        dfResult = pd.DataFrame()
        try:
            if WifiType.split('_')[-1] == "tx":
                df = pd.read_csv(filePath, header=0, sep=',', encoding='UTF-8',
                                 usecols=['Result', 'Dut_No', 'type', 'channel', 'antenna', 'bandwidth',
                                          'data_rate', 'EVM_Spec', 'EVM', 'Power_Spec', 'Power'])
                df = df[~df['Result'].isin(['TotalPASS', 'TotalPass'])].drop('Result', axis=1)

            elif WifiType.split('_')[-1] == "rx":
                df = pd.read_csv(filePath, header=0, sep=',', encoding='UTF-8',
                                 usecols=['Result', 'Dut_No', 'type', 'channel', 'antenna', 'bandwidth',
                                          'data_rate', 'RX_PER'])
                df = df[~df['Result'].isin(['TotalPASS', 'TotalPass'])].drop('Result', axis=1)
            for i, dut_no in enumerate(df['Dut_No'].unique()):
                # df = df[df['Dut_No'] == dut_no]
                for typeName, data in df.groupby('type'):
                    if typeName.split('_')[0] == "BLE":
                        # TODO: BLE manipulation
                        pass
                    elif typeName == f'{WifiType}':
                        dfResult = pd.concat([dfResult, data], axis=0)

            # ----------------------------------------------------------------------------------
            if dfResult.empty:
                print("-------------- Rx or Tx not Exist-------------- ")
                return None
            elif dfResult['type'].iloc[0].split('_')[-1] == 'rx':
                dfResult.to_csv(os.path.join(PathForRxSketch, f'{WifiType}.csv'), sep=',', encoding='UTF-8')
            elif dfResult['type'].iloc[0].split('_')[-1] == 'tx':
                dfResult.to_csv(os.path.join(PathForTxSketch, f'{WifiType}.csv'), sep=',', encoding='UTF-8')

        except NameError as nee:
            raise f'processingCSV -> {NameError(str(nee.args))}'
        except Exception as ee:
            raise Exception("fetch Ambar csv fail -->" + str(ee.args))

    @staticmethod
    def processingCsvAll(*args, **kwargs):
        WifiType = args[0]
        root = args[1]  # DirDataPolish
        PathForTxSketch = args[2]
        PathForRxSketch = args[3]
        dfResult = pd.DataFrame()
        try:
            for i, filename in enumerate(os.listdir(root)):
                if filename.endswith('csv'):
                    filepath = os.path.join(root, filename)
                    df = DataUtils.DigestPolishData(filepath, header=0)
                    # Create an empty list to store the single dataframes
                    dfList = [None] * df['Dut_No'].nunique()
                    # Get the total number of groups based on the Dut_No and
                    # Iterate over the unique values in the Dut_No column and create one dataframes
                    for j, dut_no in enumerate(df['Dut_No'].unique()):
                        dfList[j] = df[df['Dut_No'] == dut_no]
                        for typeName, data in dfList[j].groupby('type'):
                            if typeName.split('_')[0] == "BLE":
                                pass
                            elif typeName == f'{WifiType}':
                                dfResult = pd.concat([dfResult, data], axis=0)
                            else:
                                ooo = typeName
                                print('Skip TotalPass')

            # ----------------------------------------------------------------------------------
            # dfNG.to_csv(os.path.join(PathForTxSketch, r'TxFail.csv'), sep=',', encoding='UTF-8')
            if dfResult.empty:
                print("-------------- Tx not Exist-------------- ")
                return None
            elif dfResult['type'].iloc[0].split('_')[-1] == 'rx':
                dfResult.to_csv(os.path.join(PathForRxSketch, f'{WifiType}.csv'), sep=',', encoding='UTF-8')
            elif dfResult['type'].iloc[0].split('_')[-1] == 'tx':
                dfResult.to_csv(os.path.join(PathForTxSketch, f'{WifiType}.csv'), sep=',', encoding='UTF-8')

        except NameError as nee:
            raise f'processingCSV -> {NameError(str(nee.args))}'
        except Exception as ee:
            raise Exception("fetch Ambar csv fail -->" + str(ee.args))

        return df

    @staticmethod
    def processingXlsx(*args, **kwargs):
        # AmbarDataFrame = pd.DataFrame()
        # elif filename.endswith('xlsx'):
        # df = pd.read_excel(filepath, skiprows=range(1, 16), header=1, engine='openpyxl')
        # extractColumn = df.loc[:, :"Power_Avg"]
        # # hh=df.groupby("Dut_No", group_keys=True).apply(lambda k: k)
        # AmbarDataFrame = pd.concat([AmbarDataFrame, extractColumn], axis=1)
        # print(f"Data from file {i + 1}:")
        # print(AmbarDataFrame.head())
        # AmbarDataFrame.to_csv('Water.csv', encoding='utf-8', index_label=None)
        return None

    @staticmethod
    def DivideChunk(df: pd.DataFrame = None) -> pd.DataFrame():
        total_groups = df['Dut_No'].nunique()
        dfList = [None] * total_groups
        # Iterate over the unique values in the Dut_No column and create single dataframes
        for i, dut_no in enumerate(df['Dut_No'].unique()):
            dfList[i] = df[df['Dut_No'] == dut_no]

        for i, df in enumerate(dfList):
            print(f"DataFrame {i + 1}:")
            print(df.info())
            print("----------------------\n")
        # for i, j in enumerate(df.keys()):
        #     df.groupby("Dut_No", group_keys=True).apply(lambda k: k)

    @staticmethod
    def peekCpk(*args):
        """
        :LowerBound: Haven't define yet
        """
        try:
            with open('configFile.json', 'r', encoding='UTF-8') as f:
                content = f.read()
            params = json.loads(content)
            wifiType = params['WifiType']
            bw = params['BW']
            ant = params['Antenna']
            column_name = args[0]
            csvFilePath = os.path.dirname(os.path.abspath("main.py"))
            csv_files = [file for file in os.listdir(csvFilePath) if file.endswith(f'.csv')]
            filePath = os.path.join(csvFilePath, csv_files[-1])
            df = DataUtils.DigestSketchData(filePath)
            x_bar = df[f'{column_name}'].mean()
            s = df[f'{column_name}'].std()
            # Specify the upper and lower specification limits
            USL = df[f'{column_name}_Spec']
            LSL = x_bar + 2 * s
            # Calculate the CPK value
            CPK = min((USL - x_bar) / (3 * s), (x_bar - LSL) / (3 * s))
        except Exception as cpkErr:
            raise "CPK err" + str(cpkErr.args)
        return CPK

    @staticmethod
    def catchOutlier(x):
        q1 = np.percentile(x, 10)
        q3 = np.percentile(x, 90)
        outliers = q3 - q1
        return (x < q1 - 1.5 * outliers) | (x > q3 + 1.5 * outliers)

    @staticmethod
    def WifiTypeAggregate(path1, path2, sourcePath):  # -> pd.DataFrame()
        """
        Digest Two Csv to compare and plot ridgeLine
        """
        # resultDf = pd.DataFrame()
        try:
            csv1 = DataUtils.DigestSketchData(path1)
            csv2 = DataUtils.DigestSketchData(path2)
            if csv1.shape != csv2.shape:
                return None
            tw = csv1['type'][0].split('.')[0]

            # csv1['category'] = "csv1"
            # csv2['category'] = "csv2"
            # concatenate
            # resultDf = pd.concat([csv1, csv2], axis=0)
            file_create_time = datetime.fromtimestamp(os.path.getctime(path1)).strftime("%m%d_%H%M%S")
            # resultDf.to_csv(savePath + f'Comparison_{file_create_time}.csv', index=False, encoding='UTF-8')
            # merge
            csv2 = csv2.rename(columns={'EVM': 'EVM2', 'Power': 'Power2'})
            # resultDf = pd.merge(csv1[['channel', 'antenna', 'bandwidth', 'EVM', 'Power']],
            #                     csv2[['channel', 'EVM2', 'Power2']], how='outer', on='channel')
            # resultDf = csv1[['channel', 'antenna', 'bandwidth', 'EVM', 'Power']].set_index('channel').join(csv2[['channel', 'EVM2', 'Power2']].set_index('channel'), how='outer', on='channel') #, lsuffix='_left', rsuffix='_right')
            resultDf = pd.concat([csv1[['channel', 'antenna', 'bandwidth', 'EVM', 'Power']], csv2[['EVM2', 'Power2']]],
                                 axis=1)
            location = os.path.join(sourcePath, f'Comparison{tw}_{file_create_time}.csv')
            resultDf.to_csv(location, index=False, encoding='UTF-8')

        except TypeError as Te:
            raise str(Te.args)
        except RuntimeError:
            raise "Timeout ->" + str(RuntimeError.args)
        except Exception as AggErr:
            raise f'WifiTypeAggregate Err ' + str(AggErr.args)

        return resultDf

    def cols_select(self: int):
        df = pd.DataFrame()
        selected_columns = []
        for i in range(1, 4):
            col_index = i * self - 1
            selected_columns.append(df.iloc[:, col_index])
        return selected_columns

    @staticmethod
    def selectToBind():
        select3_columns = DataUtils.cols_select(3)  # df.iloc[:, 2::3]
        select4_columns = DataUtils.cols_select(4)  # df.iloc[:, 3::4]
        EVM_df = pd.DataFrame(','.join(pd.concat(select3_columns).astype(str)).split(',')).T
        Power_df = pd.DataFrame(','.join(pd.concat(select4_columns).astype(str)).split(',')).T

        draw_result_csv = pd.concat([EVM_df, Power_df], axis=1)
        draw_result_csv.columns = ['EVM', 'Power']
        # Print the new dataframe
        print(EVM_df.head())
        print('---------------------')
        print(Power_df.head())
        print('---------------------')
        print(draw_result_csv.head())

        draw_result_csv.dropna(inplace=True)
        draw_result_csv.to_csv('tommy0318', sep=',', encoding='utf-8')

    @staticmethod
    def ReadAmbarXlsx(self: int = 0, suffix: str = 'csv'):
        AmbarDataFrame = pd.DataFrame()
        # csvFile = filter(lambda s: s.endswith(suffix), )
        try:
            csvfiles = DataUtils.fetchTSB()
            for filename in enumerate(csvfiles):
                AmbarDataFrame = pd.read_excel(filename, skiprows=range(1, 16), header=1, engine='openpyxl')
                extractColumn = AmbarDataFrame.loc["Evm_Spec", "Power_Spec", "Power_Avg", "Symbol_Clk_Err_ppm"]
                # set_i = evb.loc["Evm_Avg_dB", "Power_Avg"]
            # concatenate all dataframes into a single dataframe
            # combined_dataframe = pd.concat(extractColumn, axis=1)

        except Exception as ex:
            raise Exception("read csv fail -->" + str(ex.args))
        return AmbarDataFrame


class PreManifest(DataOrder):
    def processingCSV(self, **kwargs):
        pass

    def queryResult(*args):
        df = args[0]
        csvFilePath = args[1]
        queryResult = args[2]
        if os.path.exists('configFile.json'):
            try:
                with open('configFile.json', 'r', encoding='UTF-8') as f:
                    content = f.read()
                params = json.loads(content)
                wifiType = params['WifiType']
                bw = params['BW']
                ant = params['Antenna']
                df_select = df.loc[
                    (df['type'] == f'{wifiType}') & (df['antenna'] == str(ant)) & (df['bandwidth'] == str(bw))]
                df_select.to_csv(f'{csvFilePath}/Query{wifiType}_Ant{ant}BW{bw}.csv', sep=',', encoding='UTF-8')
            except Exception as eee:
                raise "Query_{0} Err" + str(eee.args)
        else:
            # df[f'{specValue}_Spec'] = df[f'{specValue}_Spec'].astype(float)  # make different dtypes can compare
            df_select = df.loc[(df['Result'] == queryResult)]
            # df_select = df_select.loc[(df['type'] == f'{wifiType}') & (df['antenna'] == str(ant)) & (df['bandwidth'] == str(bw))]
            # df = df.loc[(df['Result'] == f'{queryResult}')]
            # df = df.loc[(df['antenna'] == str(ant))]
            # df_select = df.loc[(df['bandwidth'] == str(bw))]
            # df_select = df[(df['Result'] == f'{queryResult}') & (df['antenna'] == int(ant)) & (df['bandwidth'] == int(bw))]
            # df_select.index.name = "No."
            if df_select.empty:
                return f'No {queryResult} data or no data correspond with ConfigFile'
            else:
                df_select.to_csv(f'{csvFilePath}/Query{queryResult}.csv', sep=',', encoding='UTF-8')

        return df_select.iloc[:, 0:12].to_string(index=False)

    # def queryResult(self, **kwargs):
    #     * df['col'].quantile()
    #     * df['col'].describe()
    #     * df['col'].shape()
    #     * df['col'].value_count()
    #     * ...
    #     """
    #     try:
    #         csv_files = [file for file in os.listdir(csvFilePath) if file.endswith('.csv')]
    #         filePath = os.path.join(csvFilePath, csv_files[-1])
    #         df = DataUtils.DigestPolishData(filePath)
    #         df = df[df['Result'] == "Fail"]
    #         dfDefect = df[~(df['Result'] == 'TotalPass')].drop('Result', axis=1)
    #         q = dfDefect.info()
    #         qq = dfDefect['Evm_Avg_dB'].describe()
    #         qqq = dfDefect['Power_Avg'].describe()
    #         print(f'1. {q} \r\n 2. {qq} \r\n 3. {qqq} \r\n')
    #         print('----------')
    #     except Exception as mee:
    #         raise f'queryResult Err-> {str(mee.args)}'

    def manifestData(self, *args) -> list:
        filePath = args[0]
        base_name, extension = os.path.splitext(filePath)
        absFilePath = os.path.abspath(filePath)
        cwd = os.getcwd()
        if absFilePath.startswith(cwd):
            header = 0
            try:
                df = DataUtils.DigestPolishData(filePath, header)
                df = df[~df['Result'].isin(['TotalPASS', 'TotalPass'])].drop('Result', axis=1)
                # cleanup & trim redundant blank -----------------------
                wifiTypeArr = [x.strip() for x in df['type'].unique().astype(str) if x.strip() != '']
                antTypeArr = [x.strip() for x in df['antenna'].unique().astype(str) if x.strip() != '']
                bwTypeArr = [x.strip() for x in df['bandwidth'].unique().astype(str) if x.strip() != '']
                chTypeArr = [x.strip() for x in df['channel'].unique().astype(str) if x.strip() != '']
                OptionArr = [wifiTypeArr, antTypeArr, bwTypeArr, chTypeArr]
                return OptionArr
            except IndexError as csvFilesNotExist:
                raise f'csvFilesNotExist: {csvFilesNotExist}'
            except Exception as ex:
                raise "processing CSV err" + str(ex.args)

    def RecordFile(*args) -> json:
        global df
        result = ""
        FileSourcePath = args[0]
        DataDir = args[1]
        DataPolishDir = args[2]
        if FileSourcePath is not None:
            file_create_time = datetime.fromtimestamp(os.path.getctime(FileSourcePath)).strftime("%m%d_%H%M%S")
            base_name, extension = os.path.splitext(FileSourcePath)
            try:
                if extension == ".csv":
                    df = pd.read_csv(FileSourcePath, skiprows=range(1, 16), header=1, sep=',', encoding='UTF-8')
                    # fail_countDut = df[df['Result'] == 'Fail'].groupby('Dut_No')['Result'].count()
                    # fail_countDut = {f"Dut_{k}": v for k, v in fail_countDut.items()}
                    # failed_dut_nos = fail_countDut[fail_countDut > 0].index.tolist()
                    # whichDutFail = df.groupby('Dut_No')['Result'].apply(lambda x: (x == 'Fail').sum()).reset_index(name='TotalFail').to_dict()
                elif extension == ".xlsx":
                    wb = openpyxl.load_workbook(FileSourcePath)
                    sheet_names = wb.sheetnames[0]
                    df = pd.read_excel(FileSourcePath, sheet_name=f'{sheet_names}', header=1, skiprows=range(1, 16))
                    # fail_countDut = {f"Dut_{k}": v for k, v in fail_countDut.count().items()}
                    # json.dumps(failed_dut_nos, sort_keys=True, indent=2, ensure_ascii=False)
                    # result = df.groupby('Dut_No')['Result'].apply(lambda x: (x == 'Fail').sum()).reset_index(name='TotalFail')
                    # result = whichDutFail.rename(columns={'Dut_No': 'DutNo', 'Result': 'Failure'})
                    # fail_counts = df.groupby('Dut_No')['Result'].apply(lambda x: f"DUT{(x == 'Fail').sum()} Fail").to_json()
                else:
                    raise "Unacceptable file"
                fail_countDut = df[df['Result'] == 'Fail'].groupby('Dut_No')['Result'].count()
                failed_dut_nos = fail_countDut[fail_countDut > 0].index.tolist()
                result = ["DutNo_" + str(item) for item in failed_dut_nos]
                file_path = os.path.join(DataDir, f'{file_create_time}_WifiData.csv')
                df = df.rename({'Evm_Avg_dB': 'EVM', 'Power_Avg': 'Power', 'Evm_Spec': 'EVM_Spec'}, axis=1)
                df.to_csv(file_path, sep=',', index=False, encoding='UTF-8')
            except Exception as ee:
                raise f'fmtPrintCsvInfo err-> {str(ee.args)}'
        return result

    def fmtPrintHeader(*args):
        CsvHeaderInfo = ['created_Time', 'LotNo', 'PassYieldRate', 'SitePCName', 'Total', 'Operation_Name', 'Pass',
                         'PGM_Name', 'Fail']
        CsvMemo = []
        FileSourcePath = args[0]
        file_create_time = datetime.fromtimestamp(os.path.getctime(FileSourcePath)).strftime("%m%d_%H%M%S")
        CsvMemo.append(file_create_time)
        try:
            base_name, extension = os.path.splitext(FileSourcePath)
            if extension == ".csv":
                iterTimes = 0
                with open(FileSourcePath, 'r', newline='\n', encoding='UTF-8') as header:
                    reader = csv.reader(header)
                    for row in reader:
                        if iterTimes < 4:
                            iterTimes = iterTimes + 1
                            CsvMemo.append(row[1])
                            CsvMemo.append(row[5])
                        else:
                            break

            if extension == ".xlsx":
                iterTimes = 0
                wb = openpyxl.load_workbook(FileSourcePath)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    if iterTimes < 4:
                        iterTimes = iterTimes + 1
                        CsvMemo.append(row[1])
                        CsvMemo.append(row[5])
                    else:
                        break
            Header = dict(zip(CsvHeaderInfo, CsvMemo))
            result = json.dumps(Header, sort_keys=True, indent=2, ensure_ascii=False)
        except Exception as e:
            raise "header err" + str(e.args)
        return result

    @staticmethod
    def OverSpec(*args):
        try:
            with open('configFile.json', 'r', encoding='UTF-8') as f:
                content = f.read()
            params = json.loads(content)
            wifiType = params['WifiType']
            bw = params['BW']
            ant = params['Antenna']
            specValue = args[0]
            csvFilePath = args[1]
            csv_files = [file for file in os.listdir(csvFilePath) if file.endswith(f'{wifiType}.csv')]
            csv_files = sorted(csv_files, key=lambda x: os.path.getmtime(os.path.join(csvFilePath, x)))
            filePath = os.path.join(csvFilePath, csv_files[-1])
            df = DataUtils.DigestSketchData(filePath)
            df[f'{specValue}_Spec'] = df[f'{specValue}_Spec'].astype(float)  # make different dtypes can compare
            df_select = df.loc[(df['antenna'] == int(ant)) & (df['bandwidth'] == int(bw))]
            selected = df_select[(df_select[f'{specValue}'] > df_select[f'{specValue}_Spec'])][f'{specValue}']
            selected.index.name = "No."
            if selected.empty:
                print("The dataframe is empty")
            else:
                print("The dataframe is not empty")
                # selected.to_csv(f'OverSpec{specValue}_Ant{ant}BW{bw}.csv', sep=',', encoding='UTF-8')
            # result = df.groupby('antenna', 'bandwidth').apply(lambda v: v > df['Power_Spec'] | v > df['Evm_Spec'])
        except Exception as eee:
            raise "OverSpec_{0} Err".format(specValue) + str(eee.args)
        return selected.describe()

    # @staticmethod
    # def ReadSASP8Xlsx(suffix: str = 'csv'):
    #     csvFile = filter(lambda s: s.endswith(suffix).first(), DataUtils.fetchTSB())
    #     try:
    #         ota_conductive = pd.read_excel(csvFile, skiprows=range(1, 16), header=1)
    #         # OTACon = TsbFormatter.makeOTAConductiveTable(ota_conductive)
    #         # if isfileexist == none:
    #
    #         # if out == "OTA":
    #         #     print("Dump OTA csv")
    #         #     ota_conductive = pd.read_excel(filelst[0], skiprows=range(1, 16), header=1)
    #         #     OTACon = makeOTAConductiveTable(ota_conductive)
    #         #     OTACon.to_csv('tryOTA.csv', encoding='utf-8', index_label=None, header=0)
    #         #
    #         # elif out == "OutBand":
    #         #     print("Dump OutBand csv")
    #         #     df_outband = pd.read_excel(filelst[1], skiprows=range(1, 16), header=1)
    #         #     outBandcsv = makeOutBandTable(df_outband)
    #         #     print(outBandcsv)
    #         #     outBandcsv.to_csv('tryOutband.csv', encoding='utf-8')
    #
    #         else:
    #             print('Key in OTA or Conductive')
    #     except NameError:
    #         # TODO : Retry
    #         print("output csv ERR!")
    #
    #     except Exception:
    #         print(f"unexpected occur {Exception} {out}", out)


if __name__ == "__main__":
    try:
        WifiType = "11a_tx"
        # DataUtils.processingCSV(WifiType)
        # PreManifest.fmtPrintCsvInfo(0)
    except Exception as e:
        raise "Digesting Err" + str(e.args)
